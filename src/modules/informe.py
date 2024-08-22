import openpyxl
import os

def crear (ruta):
    """Crea el archivo excel de salida para los reportes en la ruta especificada
    :param ruta: ruta donde se guardará el archivo excel"""
    # Creamos el excel en la ruta especificada por el parametro
    wb = openpyxl.Workbook()
    # Eliminamos la hoja por defecto
    wb.remove(wb.active)
    # Creamos las hojas del excel
    wb.create_sheet('reserva_total.prn')
    wb.create_sheet('Pmax_Pgen.prn')
    wb.create_sheet('Mayor_maxima.prn')
    wb.create_sheet('Menor_optima.prn')
    wb.create_sheet('Reserva.rep')
    wb.create_sheet('Reserva.err')
    # Creamos los enacabezados de las hojas
    wb['reserva_total.prn'].append(['Escenario','Reserva Hidro','Reserva Termica','Reserva Total'])
    wb['Pmax_Pgen.prn'].append(['IBUS','NOMBRE','ID','POT_MAX[MW]','POT_GEN[MW]','MAX_GEN[MW]','RESERVA[%]','PORCENTAJE[%]','DATO','RES_OPT[[%]'])
    wb['Mayor_maxima.prn'].append(['IBUS','NOMBRE','ID','POT_MAX[MW]','POT_GEN[MW]','MAX_GEN[MW]','RESERVA[%]','PORCENTAJE[%]','DATO','RESOPT[%]'])
    wb['Menor_optima.prn'].append(['IBUS','NOMBRE','ID','POT_MAX[MW]','POT_GEN[MW]','MAX_GEN[MW]','RESERVA[%]','PORCENTAJE[%]','DATO','RESOPT[%]'])
    wb['Reserva.rep'].append(['Escenario','Reserva Hidro','Reserva Termica','Reserva Total'])
    wb['Reserva.err'].append(['Escenario','Reserva Hidro','Reserva Termica','Reserva Total'])
    # Guardamos el excel con el nombre Reserva_salida.xlsx
    ruta_completa= ruta+'/Reserva_salida1.xlsx'
    wb.save(ruta_completa)
    # Mensaje que avisa que el archivo se ha creado en la ruta especificada
    print('Archivo creado en la ruta: ',ruta_completa)
    return

def reserva_total(ruta):
    """Completa los datos de la hoja reserva_total.prn
    :param ruta: ruta donde se encuentra el archivo excel de entrada"""
    #Abrimos el archivo de excel
    workbook = openpyxl.load_workbook(ruta)
    #Seleccionamos la hoja donde vamos a completar con datos
    sheet = workbook['reserva_total.prn']
    #Escribimos los datos en la hoja debajo de los encabezados en la fila 2

    return

def Pmax_Pgen(ruta,ibus,nombre,id,pot_max,pot_gen,max_gen,reserva,porcentaje,dato,resopt):
    """Completa los datos de la hoja Pmax_Pgen.prn
    :param ruta: ruta donde se encuentra el archivo excel de entrada"""
    #Abrimos el archivo de excel
    workbook = openpyxl.load_workbook(ruta)
    #Seleccionamos la hoja donde vamos a completar con datos
    sheet = workbook['Pmax_Pgen.prn']
    #Escribimos los datos en la primer fila que encontramos vacio
    for i in range(2, sheet.max_row+1):
        if sheet.cell(row=i, column=1).value == None:
            sheet.cell(row=i, column=1).value = ibus
            sheet.cell(row=i, column=2).value = nombre
            sheet.cell(row=i, column=3).value = id
            sheet.cell(row=i, column=4).value = pot_max
            sheet.cell(row=i, column=5).value = pot_gen
            sheet.cell(row=i, column=6).value = max_gen
            sheet.cell(row=i, column=7).value = reserva
            sheet.cell(row=i, column=8).value = porcentaje
            sheet.cell(row=i, column=9).value = dato
            sheet.cell(row=i, column=10).value = resopt
            break


    return

def Mayor_maxima(ruta):
    """Completa los datos de la hoja Mayor_maxima.prn
    :param ruta: ruta donde se encuentra el archivo excel de entrada"""
    #Abrimos el archivo de excel
    workbook = openpyxl.load_workbook(ruta)
    #Seleccionamos la hoja donde vamos a completar con datos
    sheet = workbook['Mayor_maxima.prn']
    #Escribimos los datos en la hoja debajo de los encabezados en la fila 2

    return

def Reserva_rep(ruta):
    """Completa los datos de la hoja Reserva.rep
    :param ruta: ruta donde se encuentra el archivo excel de entrada"""
    #Abrimos el archivo de excel
    workbook = openpyxl.load_workbook(ruta)
    #Seleccionamos la hoja donde vamos a completar con datos
    sheet = workbook['Reserva.rep']
    #Escribimos los datos en la hoja debajo de los encabezados en la fila 2

    return

def Reserva_err(ruta):
    """Completa los datos de la hoja Reserva.err
    :param ruta: ruta donde se encuentra el archivo excel de entrada"""
    #Abrimos el archivo de excel
    workbook = openpyxl.load_workbook(ruta)
    #Seleccionamos la hoja donde vamos a completar con datos
    sheet = workbook['Reserva.err']
    #Escribimos los datos en la hoja debajo de los encabezados en la fila 2

    return
