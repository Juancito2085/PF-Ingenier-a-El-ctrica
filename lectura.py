import openpyxl

# Cargar el archivo de Excel
workbook = openpyxl.load_workbook('Reserva_entrada.xlsx')

# Determinamos las hojas de trabajo
sheet1 = workbook['Parametros']
sheet2 = workbook['Generadores SADI']
sheet3 = workbook['Regiones paises limitrofes']
sheet4 = workbook['Generadores que no suman']

parametros=list()
# Leemos la segunda columna las primeras 4 filas
for i in range(1, 5):
    cell = sheet1.cell(row=i, column=2)
    parametros.append(cell.value)

bus=list()
governor=list()
CON=list()
porcentaje=list()
idg=list()
comentario=list()
tipo=list()

i=2
while sheet2.cell(row=i, column=1).value != None: 
    bus.append(sheet2.cell(row=i, column=1).value)
    governor.append(sheet2.cell(row=i, column=2).value)
    CON.append(sheet2.cell(row=i, column=3).value)
    porcentaje.append(sheet2.cell(row=i, column=4).value)
    idg.append(sheet2.cell(row=i, column=5).value)
    comentario.append(sheet2.cell(row=i, column=6).value)
    tipo.append(sheet2.cell(row=i, column=7).value)
    i+=1

n_area=list()
comment=list()
i=2
while sheet3.cell(row=i, column=1).value != None:
    n_area.append(sheet3.cell(row=i, column=1).value)
    comment.append(sheet3.cell(row=i, column=2).value)
    i+=1

ibus=list()
nombre=list()
id=list()
i=2
while sheet4.cell(row=i, column=1).value != None:
    ibus.append(sheet4.cell(row=i, column=1).value)
    nombre.append(sheet4.cell(row=i, column=2).value)
    id.append(sheet4.cell(row=i, column=3).value)
    i+=1